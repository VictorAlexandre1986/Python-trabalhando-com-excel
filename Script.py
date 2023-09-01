from openpyxl.workbook import Workbook
from openpyxl import load_workbook


#Criando um workbook 
wb = Workbook()  

#Carregando uma planilha
lwb = load_workbook('clientes2.xlsx')

#Criando uma worksheet
ws = lwb.active

##Pegando uma celula especifica da pagina padrao
print(f'{ws["B2"].value} - Sexo: {ws["C2"].value} - Dt_Nasc: {ws["D2"].value} - Profissão: {ws["E2"].value} - R$ {ws["F2"].value}')

#Grab a whole column
column_b = ws['B']

for cell in column_b:
    print(f'{cell.value}')
    
    
#Obtendo as celulas de uma linhas especifica
row = ws['1']
for r in row:
    print(r.value, end=" - ")
print()
    
#Obtendo as celulas de uma linha especifica
range = ws['B2':'F2']
for cell in range:
    for x in cell:
        print(x.value, end=' ')
   
#----------------------------------------------------------
     
#Obtendo os workbooks disponiveis
print(wb.sheetnames)

#Criando um workbook
wb.create_sheet(title='Fornecedores', index=2)


#Obtendo a pagina
page_fornecedores = wb['Fornecedores']

#Inserindo valor as celulas
page_fornecedores.append(['Nome','CNPJ','Contato'])
page_fornecedores.append(['Gerdau','1111111111','12996541023'])

#É precisa salvar a planilha para que as informações permaneçam
wb.save(filename = 'sample_book.xlsx')

#obtendo a pagina
fornecedores_page = wb['Fornecedores']

#Imprimindo dados de cada linha
#min_row define apartir de onde deve começar a leitura e max o limite, se deixa sem ele vai ate o final
for rows in fornecedores_page.iter_rows(min_row=2,max_row=3):
    print(rows[0].value, row[1].value, rows[2].value)
    

#Atribuindo um valor a uma celula
for rows in fornecedores_page.iter_rows(min_row=2,max_row=3):
    for cell in rows:
        if (cell.value == "Gerdau"):
            cell.value = 'Embraer'
        
#É preciso salvar a modificação
wb.save(filename = 'sample_book.xlsx')        